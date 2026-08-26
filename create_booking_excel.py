import sys
import subprocess

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ---------------------------------------------------------
# TAB 1: DASHBOARD
# ---------------------------------------------------------
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.views.sheetView[0].showGridLines = True

# Title Block
ws_dash.merge_cells("A1:G2")
title_cell = ws_dash["A1"]
title_cell.value = "AbridMoroccoTrip — Booking Management Dashboard"
title_cell.font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="B84F2B", end_color="B84F2B", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# KPI Summary Cards
kpis = [
    ("Total Bookings", "=COUNTA('Bookings'!A5:A100)", "C4:D4", "C5:D5", "1F4E79"),
    ("Total Revenue ($)", "=SUM('Bookings'!I5:I100)", "E4:F4", "E5:F5", "2E7D32"),
    ("Pending Inquiries", '=COUNTIF(\'Bookings\'!K5:K100, "Pending")', "G4:H4", "G5:H5", "E65100"),
    ("Confirmed Trips", '=COUNTIF(\'Bookings\'!K5:K100, "Confirmed")', "I4:J4", "I5:J5", "0288D1")
]

for title, formula, label_range, val_range, color in kpis:
    ws_dash.merge_cells(label_range)
    l_cell = ws_dash[label_range.split(":")[0]]
    l_cell.value = title
    l_cell.font = Font(size=10, bold=True, color="FFFFFF")
    l_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    l_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_dash.merge_cells(val_range)
    v_cell = ws_dash[val_range.split(":")[0]]
    v_cell.value = formula
    v_cell.font = Font(size=18, bold=True, color=color)
    v_cell.alignment = Alignment(horizontal="center", vertical="center")
    if "$" in title:
        v_cell.number_format = "$#,##0"

# Tour Breakdown Table
ws_dash["B8"] = "Tour Performance Summary"
ws_dash["B8"].font = Font(size=14, bold=True, color="2C1A12")

headers_dash = ["Tour Package", "Base Price", "Bookings Count", "Total Revenue"]
for col_idx, text in enumerate(headers_dash, start=2):
    cell = ws_dash.cell(row=9, column=col_idx, value=text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="3F2A20", end_color="3F2A20", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

tours_list = [
    ("Marrakech Discovery Tour", 55),
    ("Atlas Mountains & Ourika Valley", 125),
    ("Essaouira Coastal Escape", 95),
    ("Merzouga Desert Adventure (3D/2N)", 145),
    ("Imperial Cities Circuit (8D/7N)", 2950),
    ("Sunrise Balloon over Marrakech", 260),
    ("Private Tailor-Made Trip", 1500)
]

for row_offset, (tour_name, price) in enumerate(tours_list, start=10):
    ws_dash.cell(row=row_offset, column=2, value=tour_name)
    p_cell = ws_dash.cell(row=row_offset, column=3, value=price)
    p_cell.number_format = "$#,##0"
    
    cnt_cell = ws_dash.cell(row=row_offset, column=4, value=f'=COUNTIF(Bookings!E$5:E$100, B{row_offset})')
    rev_cell = ws_dash.cell(row=row_offset, column=5, value=f'=SUMIF(Bookings!E$5:E$100, B{row_offset}, Bookings!I$5:I$100)')
    rev_cell.number_format = "$#,##0"

# ---------------------------------------------------------
# TAB 2: BOOKINGS (MAIN TRACKER)
# ---------------------------------------------------------
ws_book = wb.create_sheet(title="Bookings")
ws_book.views.sheetView[0].showGridLines = True

# Title
ws_book.merge_cells("A1:M2")
b_title = ws_book["A1"]
b_title.value = "AbridMoroccoTrip — Master Booking Register"
b_title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
b_title.fill = PatternFill(start_color="B84F2B", end_color="B84F2B", fill_type="solid")
b_title.alignment = Alignment(horizontal="center", vertical="center")

b_headers = [
    "Booking ID", "Date Received", "Customer Name", "Contact (Phone/Email)", 
    "Interested Tour", "Travel Date", "Travelers", "Price / Person ($)", 
    "Total Price ($)", "Payment Status", "Booking Status", "Assigned Guide", "Special Requests & Notes"
]

for col_idx, h_text in enumerate(b_headers, start=1):
    cell = ws_book.cell(row=4, column=col_idx, value=h_text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="3F2A20", end_color="3F2A20", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sample_data = [
    ("AMT-2026-001", "2026-07-20", "Jean Dupont", "+33 6 12 34 56 78 | jean@example.fr", "Merzouga Desert Adventure (3D/2N)", "2026-08-10", 2, 145, "Paid", "Confirmed", "Abdellah", "Vegetarian meals requested"),
    ("AMT-2026-002", "2026-07-21", "Sarah Jenkins", "+44 7700 900077 | sarah.j@example.co.uk", "Marrakech Discovery Tour", "2026-08-02", 4, 55, "Unpaid (Cash)", "Confirmed", "Abdellah", "Hotel pickup at Riad Alma"),
    ("AMT-2026-003", "2026-07-22", "Marco Müller", "+49 151 23456789 | marco@example.de", "Atlas Mountains & Ourika Valley", "2026-08-05", 2, 125, "Deposit Paid", "Confirmed", "Hassan", "Private 4x4 needed"),
    ("AMT-2026-004", "2026-07-24", "Maria Garcia", "+34 612 345 678 | maria@example.es", "Essaouira Coastal Escape", "2026-08-12", 3, 95, "Unpaid (Cash)", "Pending", "Abdellah", "Prefers Spanish speaking guide"),
    ("AMT-2026-005", "2026-07-25", "John Smith", "+1 212 555 0199 | john.smith@example.com", "Imperial Cities Circuit (8D/7N)", "2026-09-01", 2, 2950, "Deposit Paid", "Confirmed", "Youssef", "Luxury riads preferred"),
    ("AMT-2026-006", "2026-07-26", "Lisa Ray", "+1 312 555 0144 | lisa@example.com", "Sunrise Balloon over Marrakech", "2026-08-03", 2, 260, "Paid", "Confirmed", "Flight Partner", "Honeymoon couple - special touch"),
    ("AMT-2026-007", "2026-07-27", "Emma Watson", "+1 416 555 0188 | emma@example.ca", "Private Tailor-Made Trip", "2026-09-15", 5, 1200, "Inquired", "Pending", "Abdellah", "Custom 10-day luxury itinerary requested")
]

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

for row_idx, data in enumerate(sample_data, start=5):
    ws_book.cell(row=row_idx, column=1, value=data[0]) # ID
    ws_book.cell(row=row_idx, column=2, value=data[1]) # Date Rec
    ws_book.cell(row=row_idx, column=3, value=data[2]) # Name
    ws_book.cell(row=row_idx, column=4, value=data[3]) # Contact
    ws_book.cell(row=row_idx, column=5, value=data[4]) # Tour
    ws_book.cell(row=row_idx, column=6, value=data[5]) # Travel Date
    ws_book.cell(row=row_idx, column=7, value=data[6]) # Travelers
    
    p_cell = ws_book.cell(row=row_idx, column=8, value=data[7]) # Price / person
    p_cell.number_format = "$#,##0"
    
    tot_cell = ws_book.cell(row=row_idx, column=9, value=f"=G{row_idx}*H{row_idx}") # Total Price Formula
    tot_cell.number_format = "$#,##0"
    
    ws_book.cell(row=row_idx, column=10, value=data[8]) # Payment Status
    ws_book.cell(row=row_idx, column=11, value=data[9]) # Booking Status
    ws_book.cell(row=row_idx, column=12, value=data[10]) # Guide
    ws_book.cell(row=row_idx, column=13, value=data[11]) # Notes

    for c in range(1, 14):
        ws_book.cell(row=row_idx, column=c).border = thin_border

# Data Validation for Statuses
dv_payment = DataValidation(type="list", formula1='"Paid, Unpaid (Cash), Deposit Paid, Inquired, Refunded"', allow_blank=True)
dv_booking = DataValidation(type="list", formula1='"Pending, Confirmed, Completed, Cancelled"', allow_blank=True)
dv_tours = DataValidation(type="list", formula1='"Marrakech Discovery Tour, Atlas Mountains & Ourika Valley, Essaouira Coastal Escape, Merzouga Desert Adventure (3D/2N), Imperial Cities Circuit (8D/7N), Sunrise Balloon over Marrakech, Private Tailor-Made Trip"', allow_blank=True)

ws_book.add_data_validation(dv_payment)
ws_book.add_data_validation(dv_booking)
ws_book.add_data_validation(dv_tours)

dv_payment.add("J5:J100")
dv_booking.add("K5:K100")
dv_tours.add("E5:E100")

# Auto-fit columns
for ws in [ws_dash, ws_book]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

ws_book.column_dimensions['A'].width = 16
ws_book.column_dimensions['C'].width = 18
ws_book.column_dimensions['D'].width = 32
ws_book.column_dimensions['E'].width = 35
ws_book.column_dimensions['M'].width = 35

output_path = "C:/Users/LENOVO THINKPAD/Documents/Nouveau dossier/abridmoroccotrip-main/AbridMoroccoTrip_Booking_System.xlsx"
wb.save(output_path)
print(f"Workbook successfully created at: {output_path}")
